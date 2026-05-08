from django.http import HttpResponse
from openpyxl import Workbook

from core.utils import Styles
from organization.models import ProjectPhase, Discipline, AuthoringSoftware, LodReference, BimSpecification, BimExpert

class ExcelOrganizationExporter():
  def __init__(self):
    self.wb = Workbook()
    self.wb.remove(self.wb.active)

    self.phases = ProjectPhase.objects.all()
    self.disciplines = Discipline.objects.all()
    self.authoring_softwares = AuthoringSoftware.objects.all()
    self.lod_references = LodReference.objects.all()
    self.bim_specifications = BimSpecification.objects.all()
    self.bim_experts = BimExpert.objects.all()

  def build_organization_data(self):
    # SETUP SHEETS
    ws_fasi = self.wb.create_sheet(title='Fasi Progetto')
    ws_discipline = self.wb.create_sheet(title='Discipline')
    ws_softwares = self.wb.create_sheet(title='Software')
    ws_schede_lod = self.wb.create_sheet(title='Schede LOD')
    ws_specifiche = self.wb.create_sheet(title='Specifiche')
    ws_professionisti = self.wb.create_sheet(title='Professionisti')

    # SETUP COLUMN DIMENSIONS
    ws_fasi.column_dimensions['A'].width = 20
    ws_fasi.column_dimensions['B'].width = 60

    ws_discipline.column_dimensions['A'].width = 20
    ws_discipline.column_dimensions['B'].width = 60

    ws_softwares.column_dimensions['A'].width = 20
    ws_softwares.column_dimensions['B'].width = 60

    ws_schede_lod.column_dimensions['A'].width = 20
    ws_schede_lod.column_dimensions['B'].width = 60

    ws_specifiche.column_dimensions['A'].width = 20
    ws_specifiche.column_dimensions['B'].width = 60

    ws_professionisti.column_dimensions['A'].width = 20
    ws_professionisti.column_dimensions['B'].width = 60

    # SETUP HEADERS
    ws_fasi.append([       
      'Nome',
      'Descrizione'
      ])
    ws_fasi.row_dimensions[1].height = 20
    for cell in ws_fasi[1]:
      cell.style = Styles.header

    ws_discipline.append([       
      'Nome',
      'Descrizione'
      ])
    ws_discipline.row_dimensions[1].height = 20
    for cell in ws_discipline[1]:
      cell.style = Styles.header

    ws_softwares.append([       
      'Nome',
      'Descrizione'
      ])
    ws_softwares.row_dimensions[1].height = 20
    for cell in ws_softwares[1]:
      cell.style = Styles.header

    ws_schede_lod.append([       
      'Nome',
      'Descrizione'
      ])
    ws_schede_lod.row_dimensions[1].height = 20
    for cell in ws_schede_lod[1]:
      cell.style = Styles.header

    ws_specifiche.append([       
      'Nome',
      'Descrizione'
      ])
    ws_specifiche.row_dimensions[1].height = 20
    for cell in ws_specifiche[1]:
      cell.style = Styles.header

    ws_professionisti.append([       
      'Nome',
      'Descrizione'
      ])
    ws_professionisti.row_dimensions[1].height = 20
    for cell in ws_professionisti[1]:
      cell.style = Styles.header
    

    # SETUP CONTENT
    for count, project_phase in enumerate(self.phases):
      ws_fasi.append([         
        project_phase.name,
        project_phase.description, 
        ])
      ws_fasi.row_dimensions[ws_fasi.max_row].height = 20
      for cell in ws_fasi[ws_fasi.max_row]:
        cell.style = Styles.standard_cell

    for count, discipline in enumerate(self.disciplines):
      ws_discipline.append([         
        discipline.name,
        discipline.description, 
        ])
      ws_discipline.row_dimensions[ws_discipline.max_row].height = 20
      for cell in ws_discipline[ws_discipline.max_row]:
        cell.style = Styles.standard_cell

    for count, software in enumerate(self.authoring_softwares):
      ws_softwares.append([         
        software.name,
        software.description, 
        ])
      ws_softwares.row_dimensions[ws_softwares.max_row].height = 20
      for cell in ws_softwares[ws_softwares.max_row]:
        cell.style = Styles.standard_cell

    for count, lod_reference in enumerate(self.lod_references):
      ws_schede_lod.append([         
        lod_reference.name,
        lod_reference.description, 
        ])
      ws_schede_lod.row_dimensions[ws_schede_lod.max_row].height = 20
      for cell in ws_schede_lod[ws_schede_lod.max_row]:
        cell.style = Styles.standard_cell

    for count, specification in enumerate(self.bim_specifications):
      ws_specifiche.append([         
        specification.name,
        specification.description, 
        ])
      ws_specifiche.row_dimensions[ws_specifiche.max_row].height = 20
      for cell in ws_specifiche[ws_specifiche.max_row]:
        cell.style = Styles.standard_cell

    for count, bim_expert in enumerate(self.bim_experts):
      ws_professionisti.append([         
        bim_expert.name,
        bim_expert.description, 
        ])
      ws_professionisti.row_dimensions[ws_professionisti.max_row].height = 20
      for cell in ws_professionisti[ws_professionisti.max_row]:
        cell.style = Styles.standard_cell
    
  def export_organization_data(self):    
    self.build_organization_data()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Organization Data.xlsx"'
    self.wb.save(response)
    return response  
