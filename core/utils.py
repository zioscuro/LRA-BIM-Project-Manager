from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill

class Styles():
  thin_line = Side(border_style="thin", color="000000")

  title = NamedStyle(name="title")
  title.font = Font(bold=True)
  title.alignment = Alignment(horizontal='center')
  title.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
  title.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line)

  header = NamedStyle(name="header")
  header.font = Font(bold=True)
  header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line)

  standard_cell = NamedStyle(name="standard cell")
  standard_cell.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line)

  model_header = NamedStyle(name="model header")
  model_header.font = Font(bold=True)
  model_header.fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type = "solid")
  model_header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line) 

  info_header = NamedStyle(name="info header")
  info_header.font = Font(bold=True)
  info_header.fill = PatternFill(start_color='ffe699', end_color='ffe699', fill_type = "solid")
  info_header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line) 

  report_header = NamedStyle(name="report header")
  report_header.font = Font(bold=True)
  report_header.fill = PatternFill(start_color='f8cbad', end_color='f8cbad', fill_type = "solid")
  report_header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line) 
  report_header.alignment = Alignment(textRotation=90, wrap_text=True, horizontal='center', vertical='center')