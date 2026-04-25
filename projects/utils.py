from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from pandas import read_excel

from organization.models import AuthoringSoftware, BimExpert, BimSpecification, Discipline, LodReference
from projects.models import BimProject, BimModel, InfoSheet, Report, ClashTest, ValidationTest

class Styles():
  thin_line = Side(border_style="thin", color="000000")

  title = NamedStyle(name="title")
  title.font = Font(bold=True)
  title.alignment = Alignment(horizontal='center')
  title.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type = "solid")
  title.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line)

  header = NamedStyle(name="header")
  header.font = Font(bold=True)
  header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line)

  standard_cell = NamedStyle(name="standard cell")
  standard_cell.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line)

  model_header = NamedStyle(name="model header")
  model_header.font = Font(bold=True)
  model_header.fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type = "solid")
  model_header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line) 

  info_header = NamedStyle(name="info header")
  info_header.font = Font(bold=True)
  info_header.fill = PatternFill(start_color='ffe699', end_color='ffe699', fill_type = "solid")
  info_header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line) 

  report_header = NamedStyle(name="report header")
  report_header.font = Font(bold=True)
  report_header.fill = PatternFill(start_color='f8cbad', end_color='f8cbad', fill_type = "solid")
  report_header.border = Border(top=thin_line, left=thin_line, right=thin_line, bottom=thin_line) 
  report_header.alignment = Alignment(textRotation=90, wrap_text=True, horizontal='center', vertical='center')

class ExcelExporter():
  def __init__(self, bim_object):
    self.wb = Workbook()
    self.wb.remove(self.wb.active)  
    
    if isinstance(bim_object, BimProject):
      self.bim_project = bim_object
      self.bim_models = bim_object.bim_models.all()

    if isinstance(bim_object, BimModel):
      self.bim_model = bim_object
      self.info_sheets = bim_object.info_sheets.all()

  def build_model_register(self):
    # SETUP SHEET
    ws = self.wb.create_sheet(title='model_register')

    # SETUP COLUMN DIMENSIONS
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20

    # SETUP TITLE
    ws['A1'] = f'Registro modelli - Progetto: {self.bim_project.name}'
    ws['A1'].style = Styles.title
    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:I1')

    # SETUP HEADERS
    ws.append([
      'n.', 
      'Nome modello', 
      'Disciplina', 
      'Software',
      'Scheda LOD', 
      'Progettista', 
      'BIM Manager', 
      'BIM Coordinator', 
      'BIM Specialist'
      ])
    ws.row_dimensions[2].height = 20
    for cell in ws[2]:
      cell.style = Styles.header

    # SETUP CONTENT
    for count, bim_model in enumerate(self.bim_models):
      ws.append([
        count+1, 
        bim_model.name, 
        bim_model.discipline.name, 
        bim_model.authoringSoftware.name, 
        bim_model.lodReference.name, 
        bim_model.designer.name,
        bim_model.bim_manager.name,
        bim_model.bim_coordinator.name,
        bim_model.bim_specialist.name,
        ])
      ws.row_dimensions[ws.max_row].height = 20
      for cell in ws[ws.max_row]:
        cell.style = Styles.standard_cell
    
  def export_model_register(self):    
    self.build_model_register()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Model_Register_{self.bim_project.name}.xlsx"'
    self.wb.save(response)
    return response
  
  def build_project_info_sheets(self):
    for bim_model in self.bim_models:
      # SETUP SHEET    
      info_sheets = bim_model.info_sheets.all()
      ws = self.wb.create_sheet(title=bim_model.name)

      # SETUP COLUMN DIMENSIONS
      ws.column_dimensions['A'].width = 5
      ws.column_dimensions['B'].width = 20
      ws.column_dimensions['C'].width = 25
      ws.column_dimensions['D'].width = 15

      # SETUP TITLE
      ws['A1'] = f'Schede Informative modello: {bim_model.name}'
      ws['A1'].style = Styles.title
      ws.row_dimensions[1].height = 20
      ws.merge_cells('A1:D1')

      # SETUP HEADERS
      ws.append([
        'n.', 
        'Nome Scheda', 
        'Descrizione Scheda', 
        'Tipo Scheda'
        ])
      ws.row_dimensions[2].height = 20
      for cell in ws[2]:         
        cell.style = Styles.header

      # SETUP CONTENT
      for count, sheet in enumerate(info_sheets):
        ws.append([
          count+1, 
          sheet.name, 
          sheet.description, 
          sheet.sheet_type
          ])

        for cell in ws[ws.max_row]: 
          cell.style = Styles.standard_cell

  def export_project_info_sheets(self):

    self.build_project_info_sheets()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  
    response['Content-Disposition'] = f'attachment; filename="Project_Info_Sheets_{self.bim_project.name}.xlsx"'
    self.wb.save(response)
    return response

  def build_model_info_sheets(self):
    for sheet in self.info_sheets:
      # SETUP SHEET    
      reports = sheet.reports.all()
      ws = self.wb.create_sheet(sheet.name)

      # SETUP COLUMN DIMENSIONS
      ws.column_dimensions['A'].width = 5
      ws.column_dimensions['B'].width = 25
      ws.column_dimensions['C'].width = 50
      ws.column_dimensions['D'].width = 15
      ws.column_dimensions['E'].width = 15
      ws.column_dimensions['F'].width = 15
      ws.column_dimensions['G'].width = 15
      ws.column_dimensions['H'].width = 15

      # SETUP MODEL HEADERS
      ws['A1'] = 'DATI MODELLO'
      ws['A1'].style = Styles.model_header
      ws['A1'].alignment = Alignment(textRotation=90, wrap_text=True, horizontal='center', vertical='center')
      ws.merge_cells('A1:A5')

      ws['B1'] = 'Nome Modello'
      ws['C1'] = self.bim_model.name

      ws['B2'] = 'Disciplina'
      ws['C2'] = self.bim_model.discipline.name

      ws['B3'] = 'Progettista'
      ws['C3'] = self.bim_model.designer.name

      ws['B4'] = 'Software BIM Authoring'
      ws['C4'] = self.bim_model.authoringSoftware.name

      ws['B5'] = 'LOD di riferimento'
      ws['C5'] = self.bim_model.lodReference.name

      for row in ws.iter_cols(min_col=2, max_col=2, min_row=1, max_row=5):
        for cell in row:
          cell.style = Styles.model_header
      for row in ws.iter_cols(min_col=3, max_col=3, min_row=1, max_row=5):
        for cell in row:
          cell.style = Styles.standard_cell

      # SETUP INFO SHEETS HEADERS
      ws['A6'] = 'DATI SCHEDA'
      ws['A6'].style = Styles.info_header
      ws['A6'].alignment = Alignment(textRotation=90, wrap_text=True, horizontal='center', vertical='center')
      ws.merge_cells('A6:A8')

      ws['B6'] = 'Nome Scheda Informativa'
      ws['C6'] = sheet.name

      ws['B7'] = 'Descrizione Scheda'
      ws['C7'] = sheet.description

      ws['B8'] = 'Tipo Scheda'
      ws['C8'] = sheet.sheet_type

      for row in ws.iter_cols(min_col=2, max_col=2, min_row=6, max_row=8):
        for cell in row:
          cell.style = Styles.info_header
      for row in ws.iter_cols(min_col=3, max_col=3, min_row=6, max_row=8):
        for cell in row:
          cell.style = Styles.standard_cell
    
      ws.append([])

      # SETUP REPORT CONTENT
      for report in reports:
        if sheet.sheet_type == 'Coordination':
          tests = report.clash_tests.all()
          
          ws.append(['', 'Nome Report', report.name])
          for cell in ws[ws.max_row]:
            cell.style = Styles.standard_cell
            cell.font = Font(bold=True)
          ws.merge_cells(f'C{ws.max_row}:H{ws.max_row}')         

          ws.append(['', 'Oggetto/Specifica Report', report.description])
          for cell in ws[ws.max_row]:
            cell.style = Styles.standard_cell
          ws.merge_cells(f'C{ws.max_row}:H{ws.max_row}')
          
          ws.append([
            '', 
            'Data', 
            'Commenti', 
            'Nuove', 
            'Attive', 
            'Revisionate', 
            'Approvate', 
            'Risolte'
            ])
          for cell in ws[ws.max_row]: 
            cell.style = Styles.standard_cell          

          for test in tests:
            ws.append([
              '',
              test.date.strftime("%d/%m/%Y"), 
              test.comments, 
              test.clash_new, 
              test.clash_active, 
              test.clash_reviewed, 
              test.clash_approved, 
              test.clash_resolved
              ])            
            for cell in ws[ws.max_row]: 
              cell.style = Styles.standard_cell          
          ws.append([])         
          
        if sheet.sheet_type == 'Validation':
          tests = report.validation_tests.all()

          ws.append(['', 'Nome Report', report.name])
          for cell in ws[ws.max_row]:
            cell.style = Styles.standard_cell
            cell.font = Font(bold=True)
          ws.merge_cells(f'C{ws.max_row}:D{ws.max_row}')         

          ws.append(['', 'Descrizione Report', report.description])
          for cell in ws[ws.max_row]:
            cell.style = Styles.standard_cell
          ws.merge_cells(f'C{ws.max_row}:D{ws.max_row}')
          for cell in ws[ws.max_row]: 
            cell.style = Styles.standard_cell            
          
          ws.append(['','Data', 'Commenti', 'Difformità'])
          for cell in ws[ws.max_row]: 
            cell.style = Styles.standard_cell   

          for test in tests:
            ws.append(['', test.date.strftime("%d/%m/%Y"), test.comments, test.issues])
            for cell in ws[ws.max_row]: 
              cell.style = Styles.standard_cell       
          ws.append([])


      # SETUP REPORT HEADER
      ws['A10'] = 'ATTIVITA\' REPORT'
      ws['A10'].style = Styles.report_header
      ws.merge_cells(f'A10:A{ws.max_row}')

  def export_model_info_sheets(self):
    self.build_model_info_sheets()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  
    response['Content-Disposition'] = f'attachment; filename="{self.bim_model.name}_Info_Sheets.xlsx"'
    self.wb.save(response)  
    return response

class BimModelConfigurator():
  def __init__(self, bim_model):
    self.default_specification = get_object_or_404(BimSpecification, pk=1)
    self.bim_model = bim_model

  def default_coordination(self):
    self.bim_model.default_coordination = True
    self.bim_model.save()

    sheet_LC1 = InfoSheet(
      sheet_type = 'Coordination',
      name = 'LC1',
      description = 'default coordinamento',
      bim_model = self.bim_model
    )
    sheet_LC1.save()

    duplicates_report = Report(
      name = f'{self.bim_model.name}_duplicati',
      description = 'default report elementi duplicati',
      specification = self.default_specification,
      info_sheet = sheet_LC1,
    )
    duplicates_report.save()

    intersections_report = Report(
      name = f'{self.bim_model.name}_intersezioni',
      description = 'default report elementi intersecanti',
      specification = self.default_specification,
      info_sheet = sheet_LC1,
    )
    intersections_report.save()

  def default_validation(self):
    self.bim_model.default_validation = True
    self.bim_model.save()

    sheet_LV1 = InfoSheet(
      sheet_type ='Validation',
      name = 'LV1',
      description = 'default verifica',
      bim_model = self.bim_model
    )
    sheet_LV1.save()

    file_name_report = Report(
      name = f'{self.bim_model.name}_codifica file',
      description = 'default report nomenclatura file',
      specification = self.default_specification,
      info_sheet = sheet_LV1,
    )
    file_name_report.save()

    objects_name_report = Report(
      name = f'{self.bim_model.name}_codifica oggetti',
      description = 'default report nomenclatura oggetti nel modello',
      specification = self.default_specification,
      info_sheet = sheet_LV1,
    )
    objects_name_report.save()

class ExcelImporter():
  def __init__(self, excel_file, sheet_name, bim_project):
    self.df = read_excel(excel_file, sheet_name=sheet_name)
    self.bim_project = bim_project

  def import_model_register(self):
    for index, row in self.df.iterrows():   
      if BimModel.objects.filter(name=row['Nome modello'], description=row['Descrizione'], bim_project=self.bim_project).exists():
        break
      
      new_bim_model_discipline, created = Discipline.objects.get_or_create(name=row['Disciplina'])
      new_bim_model_discipline.save()      
      new_bim_model_software, created = AuthoringSoftware.objects.get_or_create(name=row['Software'])
      new_bim_model_software.save()      
      new_bim_model_lod, created = LodReference.objects.get_or_create(name=row['Scheda LOD'])
      new_bim_model_lod.save()      
      new_bim_model_designer, created = BimExpert.objects.get_or_create(name=row['Progettista'])
      new_bim_model_designer.save()      
      new_bim_model_bim_manager, created = BimExpert.objects.get_or_create(name=row['Bim Manager'])
      new_bim_model_bim_manager.save()      
      new_bim_model_bim_coordinator, created = BimExpert.objects.get_or_create(name=row['Bim Coordinator'])
      new_bim_model_bim_coordinator.save()      
      new_bim_model_bim_specialist, created = BimExpert.objects.get_or_create(name=row['Bim Specialist'])
      new_bim_model_bim_coordinator.save()
      
      new_bim_model = BimModel(
        name=row['Nome modello'],
        description=row['Descrizione'],
        bim_project=self.bim_project,
        discipline=new_bim_model_discipline,
        authoringSoftware=new_bim_model_software,
        lodReference=new_bim_model_lod,
        designer=new_bim_model_designer,
        bim_manager=new_bim_model_bim_manager,
        bim_coordinator=new_bim_model_bim_coordinator,
        bim_specialist=new_bim_model_bim_specialist
      )
      new_bim_model.save()

  def import_report_list(self):
    for index, row in self.df.iterrows():    
      if BimModel.objects.filter(name=row['Nome modello'], bim_project=self.bim_project).exists():
        bim_model = BimModel.objects.get(name=row['Nome modello'], bim_project=self.bim_project)
        bim_model.default_coordination = True
        bim_model.default_validation = True
        bim_model.save()

        info_sheet, created = InfoSheet.objects.get_or_create(
          name=row['Nome scheda'],
          sheet_type=row['Tipo scheda'],
          bim_model=bim_model)
        info_sheet.save()

        new_report = Report(
          name = row['Nome report'],
          description = row['Descrizione'],
          info_sheet = info_sheet,
        )
        new_report.save()

  def import_coordination_report(self):
    for bim_model in self.bim_project.bim_models.all():
      bim_model_coordination_info_sheets=bim_model.info_sheets.filter(sheet_type='Coordination')

      for info_sheet in bim_model_coordination_info_sheets:
        info_sheet_reports=info_sheet.reports.all()

        for index, row in self.df.iterrows():        
          if row['Status'] == 'Old':
            continue

          for report in info_sheet_reports:
            if report.name == row['Name']:
              new_test = ClashTest(
                clash_new=row['New'],
                clash_active=row['Active'],
                clash_reviewed=row['Reviewed'],
                clash_approved=row['Approved'],
                clash_resolved=row['Resolved'],
                report=report
              )
              new_test.save()

  def import_validation_report(self):
    for sheet, df in self.df.items():
      for bim_model in self.bim_project.bim_models.all():
        bim_model_coordination_info_sheets=bim_model.info_sheets.filter(sheet_type='Validation')

        for info_sheet in bim_model_coordination_info_sheets:
          info_sheet_reports=info_sheet.reports.all()

          for report in info_sheet_reports:
            report_issues = 0
            for index, row in df.iterrows():
              if report.name == row['Report di riferimento']:
                report_issues += 1

            if not report_issues == 0:
              new_test = ValidationTest(
                issues=report_issues,
                report=report
              )
              new_test.save()

# def handle_coordination_reports_import(clash_file, bim_project):
#   df = read_excel(clash_file, sheet_name='Clash-Results-Table')

#   for bim_model in bim_project.bim_models.all():
#     bim_model_coordination_info_sheets=bim_model.info_sheets.filter(sheet_type='Coordination')

#     for info_sheet in bim_model_coordination_info_sheets:
#       info_sheet_reports=info_sheet.reports.all()

#       for index, row in df.iterrows():        
#         if row['Status'] == 'Old':
#           continue

#         for report in info_sheet_reports:
#           if report.name == row['Name']:
#             new_test = ClashTest(
#               clash_new=row['New'],
#               clash_active=row['Active'],
#               clash_reviewed=row['Reviewed'],
#               clash_approved=row['Approved'],
#               clash_resolved=row['Resolved'],
#               report=report
#             )
#             new_test.save()

# def handle_validation_reports_import(validation_file, bim_project):
#   df_dict = read_excel(validation_file, sheet_name=None)

#   for sheet, df in df_dict.items():
#     for bim_model in bim_project.bim_models.all():
#       bim_model_coordination_info_sheets=bim_model.info_sheets.filter(sheet_type='Validation')

#       for info_sheet in bim_model_coordination_info_sheets:
#         info_sheet_reports=info_sheet.reports.all()

#         for report in info_sheet_reports:
#           report_issues = 0
#           for index, row in df.iterrows():
#             if report.name == row['Report di riferimento']:
#               report_issues += 1

#           if not report_issues == 0:
#             new_test = ValidationTest(
#               issues=report_issues,
#               report=report
#             )
#             new_test.save()

